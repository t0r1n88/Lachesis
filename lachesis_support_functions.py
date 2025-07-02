"""
Вспомогательные функции
"""
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill

class ExceedingQuantity(Exception):
    """
    Исключение для случаев когда числа уникальных значений больше 255
    """
    pass



def convert_to_int(value):
    """
    Функция для конвертации в инт
    """
    try:
        return int(value)
    except:
        return 99




def write_df_to_excel(dct_df: dict, write_index: bool) -> openpyxl.Workbook:
    """
    Функция для записи датафрейма в файл Excel
    :param dct_df: словарь где ключе это название создаваемого листа а значение датафрейм который нужно записать
    :param write_index: нужно ли записывать индекс датафрейма True or False
    :return: объект Workbook с записанными датафреймами
    """
    if len(dct_df) >= 253:
        raise ExceedingQuantity # проверяем количество значений
    wb = openpyxl.Workbook()  # создаем файл
    count_index = 0  # счетчик индексов создаваемых листов
    for name_sheet, df in dct_df.items():
        wb.create_sheet(title=name_sheet, index=count_index)  # создаем лист
        # записываем данные в лист
        none_check = None  # чекбокс для проверки наличия пустой первой строки, такое почему то иногда бывает
        for row in dataframe_to_rows(df, index=write_index, header=True):
            if len(row) == 1 and not row[0]:  # убираем пустую строку
                none_check = True
                wb[name_sheet].append(row)
            else:
                wb[name_sheet].append(row)
        if none_check:
            wb[name_sheet].delete_rows(2)

        # ширина по содержимому
        # сохраняем по ширине колонок
        for column in wb[name_sheet].columns:
            max_length = 0
            column_name = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            # для слишком длинных результатов
            if adjusted_width > 60:
                adjusted_width = 60
            wb[name_sheet].column_dimensions[column_name].width = adjusted_width
        count_index += 1

        column_number = 0  # номер колонки
        # Создаем стиль шрифта и заливки
        font = Font(color='FF000000')  # Черный цвет
        fill = PatternFill(fill_type='solid', fgColor='ffa500')  # Оранжевый цвет
        for row in wb[name_sheet].iter_rows(min_row=1, max_row=wb[name_sheet].max_row,
                                            min_col=column_number, max_col=df.shape[1] + 1):  # Перебираем строки
            if 'Итого' in str(row[column_number].value):  # делаем ячейку строковой и проверяем наличие слова Статус_
                for cell in row:  # применяем стиль если условие сработало
                    cell.font = font
                    cell.fill = fill

    return wb



def del_sheet(wb: openpyxl.Workbook, lst_name_sheet: list) -> openpyxl.Workbook:
    """
    Функция для удаления лишних листов из файла
    :param wb: объект таблицы
    :param lst_name_sheet: список удаляемых листов
    :return: объект таблицы без удаленных листов
    """
    for del_sheet in lst_name_sheet:
        if del_sheet in wb.sheetnames:
            del wb[del_sheet]

    return wb



def round_mean(value):
    """
    Функция для округления до 2 знаков
    """
    return round(value.mean(),0)

def round_mean_two(value):
    """
    Функция для округления до 2 знаков
    """
    return round(value.mean(),2)



def count_attention(row,set_value):
    """
    Функция для подсчета количества
    :param row:строка с данными
    :param set_value:значения которые нужно считать
    """
    count_val = 0
    for value in row:
        if value in set_value:
            count_val += 1
    if count_val >=2:
        return True
    else:
        return False



def sort_name_class(value):
    """
    Функция для сортировки классов вида 1А, 11Б, 2А
    :param value:
    :return:
    """
    # Разделяем строку на числовую и буквенную часть
    number_part = int(''.join(filter(str.isdigit, value)))
    letter_part = ''.join(filter(str.isalpha, value))
    return (number_part, letter_part)


def create_svod_sub(df:pd.DataFrame,lst_index:list,col_index:str,col_value:str,func:str):
    """
    Функция для создания сводов по субшкалам по всему датафрейму
    :param df: датафрейм
    :param lst_index: список с порядком строк индекса
    :param col_index: колонка по которой будет делатьться сводная таблица
    :param col_value: колонка к которой будет применятся функция
    :param func: функция свода
    :return: свод
    """
    base_svod = pd.DataFrame(index=lst_index)

    svod_level = pd.pivot_table(df, index=col_index,
                                      values=col_value,
                                      aggfunc=func)
    svod_level['% от общего'] = round(
        svod_level[col_value] / svod_level[col_value].sum(), 3) * 100


    base_svod = base_svod.join(svod_level)

    # # Создаем суммирующую строку
    base_svod.loc['Итого'] = svod_level.sum()
    base_svod.reset_index(inplace=True)
    base_svod.rename(columns={'index': 'Уровень', col_value: 'Количество'},
                           inplace=True)
    return base_svod



def create_union_svod(base_df:pd.DataFrame,dct_svod_integral:dict,dct_rename_svod_integral:dict,lst_integral:list):
    """
    Функция для создания объединенного свода по шкалам с одинаковыми названиями уровней
    :param base_df: датафрейм с результатам подсчетов
    :param dct_svod_integral: словарь где ключ это название колонки с значением шкалы а значение это название колонки с уровнем значения шкалы
    :param dct_rename_svod_integral: словарь для переименования колонок ключ это значение шкалы а значение это то как будет называться колонка в своде
    :param lst_integral:  список уровней
    :return: датафрейм
    """
    # общий датафрейм
    base_svod_df = pd.DataFrame(
        index=lst_integral)

    for key,value in dct_svod_integral.items():
        svod_level_df = pd.pivot_table(base_df, index=value,
                                       values=key,
                                       aggfunc='count')

        svod_level_df[f'{dct_rename_svod_integral[key]} % от общего'] = round(
            svod_level_df[key] / svod_level_df[
                key].sum(), 3) * 100

        base_svod_df = base_svod_df.join(svod_level_df)

        # # Создаем суммирующую строку
    base_svod_df.loc['Итого'] = base_svod_df.sum()
    base_svod_df.reset_index(inplace=True)
    # Переименовываем
    base_svod_df.rename(columns=dct_rename_svod_integral,inplace=True)
    base_svod_df.rename(columns={'index': 'Уровень'},inplace=True)

    return base_svod_df


def create_list_on_level(base_df:pd.DataFrame,out_dct:dict,lst_level:list,dct_prefix:dict):
    """
    Функция для создания списков по уровням шкал
    :param base_df: датафрейм с результатами
    :param out_dct: словарь с датафреймами
    :param lst_level: список уровней по которым нужно сделать списки
    :param dct_prefix: префиксы для названий листов
    :return: обновлейнный out dct
    """
    for key,value in dct_prefix.items():
        dct_level = dict()
        for level in lst_level:
            temp_df = base_df[base_df[key] == level]
            if temp_df.shape[0] != 0:
                dct_level[f'{dct_prefix[key]}. {level}'] = temp_df
        out_dct.update(dct_level)

    return out_dct








